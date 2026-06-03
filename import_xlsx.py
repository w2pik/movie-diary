"""
从 xlsx 导入电影数据到 SQLite 数据库
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from database import init_db, insert_movie, get_all_movies, DB_PATH

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "19e8905bb13_a6b.xlsx")


def import_data():
    """从 xlsx 导入数据"""
    if not os.path.exists(XLSX_PATH):
        print(f"❌ 找不到文件: {XLSX_PATH}")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    print(f"📂 工作表: {ws.title}, 行数: {ws.max_row}, 列数: {ws.max_column}")

    # 读取表头映射
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"📋 列: {headers}")

    # 列名 → 数据库字段映射
    col_map = {
        "电影名": "title",
        "原名称": "original_title",
        "评分": "my_rating",
        "观影时间": "watch_time",
        "电影日记": "diary",
        "导演": "director",
        "上映日期": "release_date",
        "年份": "year",
        "国家": "country",
        "豆瓣评分": "douban_rating",
        "主演": "cast_info",
        "简介": "summary",
        "片长": "duration",
        "语言": "language",
        "影片类型": "genre",
        "编剧": "writer",
        "标签": "tags",
    }

    # 映射 header index → db field
    field_map = {}
    for i, h in enumerate(headers):
        if h in col_map:
            field_map[i] = col_map[h]

    imported = 0
    skipped = 0

    for r in range(2, ws.max_row + 1):
        data = {}
        for col_idx, field in field_map.items():
            val = ws.cell(row=r, column=col_idx + 1).value
            data[field] = str(val).strip() if val is not None else ""

        title = data.get("title", "")
        if not title or title == "电影名":
            skipped += 1
            continue

        # 确保评分是数字
        try:
            data["my_rating"] = float(data.get("my_rating", 0) or 0)
        except ValueError:
            data["my_rating"] = 0

        insert_movie(data)
        imported += 1

        if imported % 50 == 0:
            print(f"  已导入 {imported} 部...")

    print(f"\n✅ 导入完成: {imported} 部电影, 跳过 {skipped} 行")


if __name__ == "__main__":
    # 如果数据库已存在，先检查
    if os.path.exists(DB_PATH):
        from database import get_db
        conn = get_db()
        count = conn.execute("SELECT COUNT(*) as c FROM movies").fetchone()["c"]
        conn.close()
        if count > 0:
            ans = input(f"数据库已有 {count} 条记录，是否覆盖？(y/N): ").strip().lower()
            if ans == 'y':
                os.remove(DB_PATH)
                print("🗑️  已删除旧数据库")
            else:
                print("❌ 已取消导入")
                sys.exit(0)

    print("🔧 初始化数据库...")
    init_db()
    print("📥 开始导入数据...")
    import_data()

    # 验证
    movies = get_all_movies(limit=3)
    print(f"\n📊 数据库共 {len(get_all_movies())} 部电影")
    print("前3部:")
    for m in movies:
        print(f"  🎬 {m['title']} ({m['year']}) — ⭐{m['douban_rating']}")
